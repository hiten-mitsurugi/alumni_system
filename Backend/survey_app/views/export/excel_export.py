"""
Excel Export Functionality
===========================
Dynamic export of complete alumni data including user info, profiles, and survey responses.
"""

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.apps import apps

from ...models import SurveyQuestion, SurveyResponse
from ...permissions import IsSurveyAdmin


@api_view(['POST'])
@permission_classes([IsSurveyAdmin])
def survey_export_view(request):
    """
    📊 SIMPLIFIED EXPORT - Exports only essential user info (Full Name, Email, Year Graduated, Program) 
    plus all survey answers for the specified form.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    import json
    from datetime import datetime
    
    try:
        # Get parameters from request body
        export_format = request.data.get('format', 'xlsx')
        category_id = request.data.get('category_id')  # Single category (backward compatibility)
        category_ids = request.data.get('category_ids', [])  # Multiple categories (new)
        date_from = request.data.get('date_from')
        date_to = request.data.get('date_to')
        
        # Support both single category_id and multiple category_ids
        if category_id and not category_ids:
            category_ids = [category_id]
        
        print(f"🔍 Starting export with filters: categories={category_ids}, date_from={date_from}, date_to={date_to}")
        
        # ===== STEP 1: Get ALL survey questions for selected categories =====
        all_questions = SurveyQuestion.objects.all().select_related('category').order_by('category__name', 'order')
        if category_ids:
            all_questions = all_questions.filter(category_id__in=category_ids)
        
        print(f"📊 Found {all_questions.count()} questions to include")
        
        # ===== STEP 2: Get ALL survey responses =====
        all_responses = SurveyResponse.objects.select_related('user', 'question', 'question__category')
        
        # Apply date filters
        if date_from:
            all_responses = all_responses.filter(submitted_at__gte=date_from)
        if date_to:
            all_responses = all_responses.filter(submitted_at__lte=date_to)
        
        # Apply category filter
        if category_ids:
            all_responses = all_responses.filter(question__category_id__in=category_ids)
        
        print(f"📝 Found {all_responses.count()} responses to include")
        
        # ===== STEP 3: Get users who have responses =====
        from django.contrib.auth import get_user_model
        User = get_user_model()

        user_ids = all_responses.values_list('user_id', flat=True).distinct()
        users = User.objects.filter(id__in=user_ids)
        
        print(f"👥 Found {users.count()} users with responses")
        
        if not users.exists():
            return Response({'error': 'No survey responses found for the specified criteria.'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # ===== STEP 4: Create workbook =====
        wb = Workbook()
        
        # ===== SHEET 1: SURVEY RESPONSES DATA =====
        ws_main = wb.active
        ws_main.title = "Survey Responses"
        
        # Build simplified headers - ONLY essential user info + survey questions
        headers = ['Full Name', 'Email', 'Year Graduated', 'Program']
        header_mapping = {}
        col_index = 0
        
        # Map first 4 columns to user fields
        header_mapping[0] = ('user', 'full_name')
        header_mapping[1] = ('user', 'email')
        header_mapping[2] = ('user', 'year_graduated')
        header_mapping[3] = ('user', 'program')
        col_index = 4
        
        # Add ALL Survey Questions
        question_columns = {}
        for question in all_questions:
            question_header = f"[{question.category.name}] {question.question_text}"
            if len(question_header) > 100:
                question_header = question_header[:97] + "..."
            headers.append(question_header)
            header_mapping[col_index] = ('question', question.id)
            question_columns[question.id] = col_index
            col_index += 1
        
        print(f"📋 Created {len(headers)} columns (4 user fields + {len(all_questions)} questions)")
        
        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            
            # Different colors for user info vs survey questions
            if col <= 4:  # User info columns
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            else:  # Survey question columns
                cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
        
        # ===== STEP 5: Get ALL responses organized by user and question =====
        user_responses = {}
        for response in all_responses:
            if response.user_id not in user_responses:
                user_responses[response.user_id] = {}
            user_responses[response.user_id][response.question_id] = response.response_data
        
        # ===== STEP 6: Write data for each user =====
        for row, user in enumerate(users, 2):
            for col_idx, (data_type, field_name) in header_mapping.items():
                col = col_idx + 1
                cell_value = ""
                
                try:
                    if data_type == 'user':
                        # Get value from User model
                        if field_name == 'full_name':
                            # Combine first_name, middle_name, last_name
                            parts = []
                            if user.first_name:
                                parts.append(user.first_name)
                            if hasattr(user, 'middle_name') and user.middle_name:
                                parts.append(user.middle_name)
                            if user.last_name:
                                parts.append(user.last_name)
                            cell_value = ' '.join(parts) if parts else ''
                        else:
                            cell_value = getattr(user, field_name, '')
                        
                    elif data_type == 'question':
                        # Get survey response
                        question_id = field_name
                        if user.id in user_responses and question_id in user_responses[user.id]:
                            response_data = user_responses[user.id][question_id]
                            
                            # Handle different response formats
                            if isinstance(response_data, dict):
                                if 'value' in response_data:
                                    cell_value = response_data['value']
                                elif 'selected_options' in response_data:
                                    cell_value = ', '.join(map(str, response_data['selected_options']))
                                elif 'rating' in response_data:
                                    cell_value = response_data['rating']
                                elif 'text' in response_data:
                                    cell_value = response_data['text']
                                elif 'answer' in response_data:
                                    cell_value = response_data['answer']
                                else:
                                    cell_value = json.dumps(response_data) if response_data else ''
                            elif isinstance(response_data, list):
                                cell_value = ', '.join(map(str, response_data))
                            else:
                                cell_value = str(response_data) if response_data else ''
                    
                    # Convert to string and handle None values
                    if cell_value is None:
                        cell_value = ''
                    elif isinstance(cell_value, (list, dict)):
                        cell_value = json.dumps(cell_value)
                    else:
                        cell_value = str(cell_value)
                    
                except Exception as e:
                    cell_value = f"Error: {str(e)}"
                
                # Write cell value
                ws_main.cell(row=row, column=col, value=cell_value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            if col <= 4:
                ws_main.column_dimensions[column_letter].width = 25  # Wider for user info
            else:
                ws_main.column_dimensions[column_letter].width = 30  # Extra wide for survey answers
        
        # ===== SHEET 2: QUESTIONS REFERENCE =====
        ws_questions = wb.create_sheet("Questions Reference")
        q_headers = ['Category', 'Question ID', 'Question Text', 'Question Type', 'Options/Scale', 'Is Required', 'Order', 'Response Count']
        
        # Write question headers
        for col, header in enumerate(q_headers, 1):
            cell = ws_questions.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Write question data
        for row, question in enumerate(all_questions, 2):
            response_count = all_responses.filter(question=question).count()
            
            # Get options/scale info
            options_info = ""
            if hasattr(question, 'options') and question.options:
                try:
                    options_data = json.loads(question.options) if isinstance(question.options, str) else question.options
                    if isinstance(options_data, list):
                        options_info = ', '.join(options_data)
                    elif isinstance(options_data, dict):
                        options_info = json.dumps(options_data)
                except:
                    options_info = str(question.options)
            
            question_data = [
                question.category.name,
                question.id,
                question.question_text,
                question.question_type,
                options_info,
                'Yes' if question.is_required else 'No',
                question.order,
                response_count
            ]
            
            for col, data in enumerate(question_data, 1):
                ws_questions.cell(row=row, column=col, value=str(data))
        
        # Auto-adjust columns for questions sheet
        for col in range(1, len(q_headers) + 1):
            ws_questions.column_dimensions[get_column_letter(col)].width = 20
        
        # ===== SHEET 3: EXPORT SUMMARY =====
        ws_summary = wb.create_sheet("Export Summary")
        summary_data = [
            ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Respondents', users.count()],
            ['Total Questions Included', all_questions.count()],
            ['Total Responses Included', all_responses.count()],
            ['Category Filter', ', '.join(map(str, category_ids)) if category_ids else 'All Categories'],
            ['Date From Filter', date_from or 'No Limit'],
            ['Date To Filter', date_to or 'No Limit'],
            ['User Info Fields', 4],  # Full Name, Email, Year Graduated, Program
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=str(value))
        
        # Auto-adjust columns for summary sheet
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 40
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'survey_responses_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print(f"✅ Export completed successfully: {filename}")
        return response
        
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': f'Export failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
