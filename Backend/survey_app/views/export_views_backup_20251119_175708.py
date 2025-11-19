"""
Survey App - Export Views
==========================
Functions for exporting survey data in various formats (Excel, PDF).
Includes comprehensive export functionality with charts and analytics.
"""

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.utils import timezone
from django.apps import apps

from ..models import SurveyCategory, SurveyQuestion, SurveyResponse, SurveyTemplate
from ..permissions import IsSurveyAdmin, IsSuperAdminOnly


# =============================================================================
# EXCEL EXPORT
# =============================================================================

@api_view(['POST'])
@permission_classes([IsSurveyAdmin])
def survey_export_view(request):
    """
    🚀 DYNAMIC EXPORT - Automatically includes ALL survey questions, user fields, and profile data.
    Adapts to any changes in questions or user models without code changes.
    Includes COMPLETE alumni directory information for comprehensive tracing.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    import json
    from datetime import datetime
    
    try:
        # Get parameters from request body
        export_format = request.data.get('format', 'xlsx')
        category_id = request.data.get('category_id')  # Single category (backward compatibility)
        category_ids = request.data.get('category_ids', [])  # Multiple categories (new)
        date_from = request.data.get('date_from')
        date_to = request.data.get('date_to')
        
        # Support both single category_id and multiple category_ids
        if category_id and not category_ids:
            category_ids = [category_id]
        
        print(f"🔍 Starting DYNAMIC export with filters: categories={category_ids}, date_from={date_from}, date_to={date_to}")
        
        # ===== STEP 1: Get ALL survey questions dynamically (not hardcoded) =====
        all_questions = SurveyQuestion.objects.all().select_related('category').order_by('category__name', 'order')
        if category_ids:
            all_questions = all_questions.filter(category_id__in=category_ids)
        
        print(f"📊 Found {all_questions.count()} questions to include")
        
        # ===== STEP 2: Get ALL survey responses dynamically =====
        all_responses = SurveyResponse.objects.select_related('user', 'question', 'question__category')
        
        # Apply date filters
        if date_from:
            all_responses = all_responses.filter(submitted_at__gte=date_from)
        if date_to:
            all_responses = all_responses.filter(submitted_at__lte=date_to)
        
        # Apply category filter
        if category_ids:
            all_responses = all_responses.filter(question__category_id__in=category_ids)
        
        print(f"📝 Found {all_responses.count()} responses to include")
        
        # ===== STEP 3: Get ALL users who have ANY responses =====
        from django.contrib.auth import get_user_model
        User = get_user_model()

        user_ids = all_responses.values_list('user_id', flat=True).distinct()
        users = User.objects.filter(id__in=user_ids).select_related('profile').prefetch_related('normalized_addresses')
        
        print(f"👥 Found {users.count()} users with responses")
        
        if not users.exists():
            return Response({'error': 'No survey responses found for the specified criteria.'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # ===== STEP 4: DYNAMICALLY get ALL fields from ALL models =====
        
        # DYNAMICALLY get ALL fields from User model
        user_model_fields = []
        user_excluded = ['password', 'last_login', 'is_superuser', 'is_staff', 'groups', 'user_permissions']
        for field in User._meta.fields:
            if field.name not in user_excluded:
                user_model_fields.append({
                    'name': field.name,
                    'verbose_name': field.verbose_name or field.name.replace('_', ' ').title()
                })
        
        # DYNAMICALLY get ALL fields from Profile model
        profile_model_fields = []
        try:
            Profile = apps.get_model('auth_app', 'Profile')
            for field in Profile._meta.fields:
                if field.name not in ['id', 'user']:
                    profile_model_fields.append({
                        'name': field.name,
                        'verbose_name': field.verbose_name or field.name.replace('_', ' ').title()
                    })
        except:
            print("⚠️ Profile model not found")
        
        # DYNAMICALLY get ALL fields from Address model
        address_model_fields = []
        try:
            Address = apps.get_model('auth_app', 'Address')
            for field in Address._meta.fields:
                if field.name not in ['id', 'user']:
                    address_model_fields.append({
                        'name': field.name,
                        'verbose_name': field.verbose_name or field.name.replace('_', ' ').title()
                    })
        except:
            print("⚠️ Address model not found")
        
        # ===== STEP 5: Create workbook =====
        wb = Workbook()
        
        # ===== SHEET 1: COMPLETE ALUMNI DATA =====
        ws_main = wb.active
        ws_main.title = "Complete Alumni Data"
        
        # Build DYNAMIC headers
        headers = []
        header_mapping = {}  # Track what each column represents
        col_index = 0
        
        # Add ALL User model fields
        for field_info in user_model_fields:
            headers.append(f"USER: {field_info['verbose_name']}")
            header_mapping[col_index] = ('user', field_info['name'])
            col_index += 1
        
        # Add ALL Profile model fields
        for field_info in profile_model_fields:
            headers.append(f"PROFILE: {field_info['verbose_name']}")
            header_mapping[col_index] = ('profile', field_info['name'])
            col_index += 1
        
        # Add ALL Address model fields (Present Address)
        for field_info in address_model_fields:
            headers.append(f"PRESENT ADDRESS: {field_info['verbose_name']}")
            header_mapping[col_index] = ('address_present', field_info['name'])
            col_index += 1
        
        # Add ALL Address model fields (Permanent Address)
        for field_info in address_model_fields:
            headers.append(f"PERMANENT ADDRESS: {field_info['verbose_name']}")
            header_mapping[col_index] = ('address_permanent', field_info['name'])
            col_index += 1
        
        # Add ALL Survey Questions DYNAMICALLY
        question_columns = {}
        for question in all_questions:
            question_header = f"SURVEY: [{question.category.name}] {question.question_text}"
            if len(question_header) > 100:
                question_header = question_header[:97] + "..."
            headers.append(question_header)
            header_mapping[col_index] = ('question', question.id)
            question_columns[question.id] = col_index
            col_index += 1
        
        print(f"📋 Created {len(headers)} dynamic columns")
        
        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            
            # Different colors for different sections
            if header.startswith("USER:"):
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            elif header.startswith("PROFILE:"):
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            elif "ADDRESS:" in header:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            elif header.startswith("SURVEY:"):
                cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
        
        # ===== STEP 6: Get ALL responses organized by user and question =====
        user_responses = {}
        for response in all_responses:
            if response.user_id not in user_responses:
                user_responses[response.user_id] = {}
            user_responses[response.user_id][response.question_id] = response.response_data
        
        # ===== STEP 7: Write data for each user DYNAMICALLY =====
        for row, user in enumerate(users, 2):
            for col_idx, (data_type, field_name) in header_mapping.items():
                col = col_idx + 1
                cell_value = ""
                
                try:
                    if data_type == 'user':
                        # Get value from User model
                        cell_value = getattr(user, field_name, '')
                        
                    elif data_type == 'profile':
                        # Get value from Profile model
                        if hasattr(user, 'profile') and user.profile:
                            cell_value = getattr(user.profile, field_name, '')
                        
                    elif data_type == 'address_present':
                        # Get value from Present Address
                        present_address = user.normalized_addresses.filter(address_category='present').first()
                        if present_address:
                            cell_value = getattr(present_address, field_name, '')
                        
                    elif data_type == 'address_permanent':
                        # Get value from Permanent Address
                        permanent_address = user.normalized_addresses.filter(address_category='permanent').first()
                        if permanent_address:
                            cell_value = getattr(permanent_address, field_name, '')
                        
                    elif data_type == 'question':
                        # Get survey response
                        question_id = field_name
                        if user.id in user_responses and question_id in user_responses[user.id]:
                            response_data = user_responses[user.id][question_id]
                            
                            # DYNAMICALLY handle ALL response formats
                            if isinstance(response_data, dict):
                                if 'value' in response_data:
                                    cell_value = response_data['value']
                                elif 'selected_options' in response_data:
                                    cell_value = ', '.join(map(str, response_data['selected_options']))
                                elif 'rating' in response_data:
                                    cell_value = response_data['rating']
                                elif 'text' in response_data:
                                    cell_value = response_data['text']
                                elif 'answer' in response_data:
                                    cell_value = response_data['answer']
                                else:
                                    cell_value = json.dumps(response_data) if response_data else ''
                            elif isinstance(response_data, list):
                                cell_value = ', '.join(map(str, response_data))
                            else:
                                cell_value = str(response_data) if response_data else ''
                    
                    # Convert to string and handle None values
                    if cell_value is None:
                        cell_value = ''
                    elif isinstance(cell_value, (list, dict)):
                        cell_value = json.dumps(cell_value)
                    else:
                        cell_value = str(cell_value)
                    
                except Exception as e:
                    cell_value = f"Error: {str(e)}"
                
                # Write cell value
                ws_main.cell(row=row, column=col, value=cell_value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws_main.column_dimensions[column_letter].width = 20
        
        # ===== SHEET 2: QUESTIONS REFERENCE =====
        ws_questions = wb.create_sheet("Questions Reference")
        q_headers = ['Category', 'Question ID', 'Question Text', 'Question Type', 'Options/Scale', 'Is Required', 'Order', 'Response Count']
        
        # Write question headers
        for col, header in enumerate(q_headers, 1):
            cell = ws_questions.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Write question data
        for row, question in enumerate(all_questions, 2):
            response_count = all_responses.filter(question=question).count()
            
            # Get options/scale info
            options_info = ""
            if hasattr(question, 'options') and question.options:
                try:
                    options_data = json.loads(question.options) if isinstance(question.options, str) else question.options
                    if isinstance(options_data, list):
                        options_info = ', '.join(options_data)
                    elif isinstance(options_data, dict):
                        options_info = json.dumps(options_data)
                except:
                    options_info = str(question.options)
            
            question_data = [
                question.category.name,
                question.id,
                question.question_text,
                question.question_type,
                options_info,
                'Yes' if question.is_required else 'No',
                question.order,
                response_count
            ]
            
            for col, data in enumerate(question_data, 1):
                ws_questions.cell(row=row, column=col, value=str(data))
        
        # ===== SHEET 3: EXPORT SUMMARY =====
        ws_summary = wb.create_sheet("Export Summary")
        summary_data = [
            ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Users Exported', users.count()],
            ['Total Questions Included', all_questions.count()],
            ['Total Responses Included', all_responses.count()],
            ['Category Filter', category_id or 'All Categories'],
            ['Date From Filter', date_from or 'No Limit'],
            ['Date To Filter', date_to or 'No Limit'],
            ['User Model Fields', len(user_model_fields)],
            ['Profile Model Fields', len(profile_model_fields)],
            ['Address Model Fields', len(address_model_fields) * 2],  # Present + Permanent
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=str(value))
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'complete_alumni_data_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print(f"✅ DYNAMIC export completed successfully: {filename}")
        return response
        
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': f'Export failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# PDF EXPORTS
# =============================================================================

def _extract_value_for_pdf(response_data):
    """Helper function to extract value from response_data for PDF export"""
    if not response_data:
        return None
    if isinstance(response_data, dict):
        for key in ['value', 'answer', 'text', 'rating', 'selected_options']:
            if key in response_data:
                return response_data[key]
        return response_data
    return response_data


@api_view(['POST'])
@permission_classes([IsSurveyAdmin])
def category_analytics_pdf_export(request):
    """
    Export category analytics as comprehensive PDF report with charts, graphs and statistics.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.lib.enums import TA_CENTER
        from datetime import datetime
        import io
        
        category_id = request.data.get('category_id')
        
        if not category_id:
            return Response({'error': 'category_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            category = SurveyCategory.objects.get(id=category_id)
        except SurveyCategory.DoesNotExist:
            return Response({'error': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get analytics data
        questions = SurveyQuestion.objects.filter(
            category=category,
            is_active=True
        ).order_by('order', 'question_text')
        
        responses = SurveyResponse.objects.filter(
            question__category=category
        ).select_related('user', 'question')
        
        from django.contrib.auth import get_user_model
        User = get_user_model()
        total_alumni = User.objects.filter(user_type=3, is_approved=True).count()
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter, 
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch,
            leftMargin=0.75*inch,
            rightMargin=0.75*inch
        )
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=26,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=18,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            spaceBefore=16,
            fontName='Helvetica-Bold'
        )
        
        question_heading_style = ParagraphStyle(
            'QuestionHeading',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=8,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )
        
        # Header Section
        story.append(Paragraph("Survey Analytics Report", title_style))
        story.append(Paragraph(f"{category.name}", subtitle_style))
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Summary Section
        story.append(Paragraph("Executive Summary", heading_style))
        
        unique_respondents = responses.values('user').distinct().count()
        response_rate = round((unique_respondents / total_alumni * 100) if total_alumni > 0 else 0, 2)
        
        summary_data = [
            ['Metric', 'Value'],
            ['Category', category.name],
            ['Report Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Questions', str(questions.count())],
            ['Total Responses Received', str(responses.count())],
            ['Unique Respondents', str(unique_respondents)],
            ['Total Alumni Population', str(total_alumni)],
            ['Overall Response Rate', f"{response_rate}%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#eff6ff')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f9ff')]),
            ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#93c5fd')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.4*inch))
        
        # Question Analytics Section
        story.append(Paragraph("Detailed Question Analytics", heading_style))
        story.append(Spacer(1, 0.15*inch))
        
        for idx, question in enumerate(questions, 1):
            question_responses = responses.filter(question=question)
            response_count = question_responses.count()
            
            # Question header
            question_text = f"{idx}. {question.question_text}"
            story.append(Paragraph(question_text, question_heading_style))
            
            # Question metadata
            meta_text = f"<i>Type: {question.get_question_type_display()} | Responses: {response_count}"
            if response_count > 0:
                q_response_rate = round((response_count / total_alumni * 100) if total_alumni > 0 else 0, 1)
                meta_text += f" | Response Rate: {q_response_rate}%"
            meta_text += "</i>"
            story.append(Paragraph(meta_text, styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
            
            # Type-specific analytics with charts (simplified for brevity)
            if question.question_type in ['text', 'textarea', 'email']:
                count_data = [[f'{response_count} text responses received']]
                count_table = Table(count_data, colWidths=[5.7*inch])
                count_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 11),
                    ('TOPPADDING', (0, 0), (-1, -1), 15),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#60a5fa')),
                ]))
                story.append(count_table)
                story.append(Paragraph("<i>Individual responses are protected for privacy</i>", styles['Normal']))
            
            story.append(Spacer(1, 0.25*inch))
            
            # Add page break every 3 questions
            if idx % 3 == 0 and idx < len(questions):
                story.append(PageBreak())
        
        # Build PDF
        doc.build(story)
        
        # Return response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_category_name = ''.join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in category.name)
        filename = f'survey_analytics_{safe_category_name}_{timestamp}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError:
        return Response({
            'error': 'ReportLab library not installed. Run: pip install reportlab'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': f'PDF export failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsSurveyAdmin])
def form_analytics_pdf_export(request):
    """
    Export complete form analytics (all categories) as comprehensive PDF report with charts, graphs and statistics.
    Accepts category_ids array to specify which categories to include.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.lib.enums import TA_CENTER
        from datetime import datetime
        import io
        
        # Get category_ids from request
        category_ids = request.data.get('category_ids', [])
        
        # Try to find the template/form name from categories
        form_title = "Survey Analytics Report"
        if category_ids:
            # Find template that contains these categories
            template = SurveyTemplate.objects.filter(
                categories__id__in=category_ids,
                is_active=True
            ).first()
            if template:
                form_title = template.name
            
            categories = SurveyCategory.objects.filter(
                id__in=category_ids,
                is_active=True
            ).order_by('order', 'created_at')
        else:
            # Default to registration survey categories
            template = SurveyTemplate.objects.filter(is_default=True, is_active=True).first()
            if template:
                form_title = template.name
            
            categories = SurveyCategory.objects.filter(
                include_in_registration=True,
                is_active=True
            ).order_by('order', 'created_at')
        
        if not categories.exists():
            return Response({'error': 'No categories found for export'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get all questions from selected categories
        category_ids = [cat.id for cat in categories]
        questions = SurveyQuestion.objects.filter(
            category_id__in=category_ids,
            is_active=True
        ).select_related('category').order_by('category__order', 'order', 'question_text')
        
        # Get all responses for these questions
        responses = SurveyResponse.objects.filter(
            question__category_id__in=category_ids
        ).select_related('user', 'question')
        
        from django.contrib.auth import get_user_model
        User = get_user_model()
        total_alumni = User.objects.filter(user_type=3, is_approved=True).count()
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter, 
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch,
            leftMargin=0.75*inch,
            rightMargin=0.75*inch
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=26,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceBefore=6,
            spaceAfter=12,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        section_heading_style = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#7c3aed'),
            spaceAfter=10,
            spaceBefore=15,
            fontName='Helvetica-Bold'
        )
        question_heading_style = ParagraphStyle(
            'QuestionHeading',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )
        
        # Build story
        story = []
        
        # Title - Use the form name
        story.append(Paragraph(form_title, title_style))
        story.append(Paragraph(f"Complete Analytics Report", subtitle_style))
        timestamp = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        story.append(Paragraph(f"<i>Generated on {timestamp}</i>", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Executive Summary
        total_responses = responses.values('user').distinct().count()
        response_rate = round((total_responses / total_alumni * 100) if total_alumni > 0 else 0, 1)
        total_questions = questions.count()
        total_categories = categories.count()
        
        story.append(Paragraph("Executive Summary", heading_style))
        summary_data = [
            ['Metric', 'Value'],
            ['Total Categories', str(total_categories)],
            ['Total Questions', str(total_questions)],
            ['Total Responses', str(total_responses)],
            ['Total Alumni', str(total_alumni)],
            ['Overall Response Rate', f"{response_rate}%"],
            ['Report Date', datetime.now().strftime('%Y-%m-%d')],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2.7*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#eff6ff')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#60a5fa')),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        story.append(summary_table)
        story.append(PageBreak())
        
        # Process each category and its questions
        for cat_idx, category in enumerate(categories, 1):
            # Category header
            story.append(Paragraph(f"Category {cat_idx}: {category.name}", section_heading_style))
            if category.description:
                story.append(Paragraph(f"<i>{category.description}</i>", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            # Get questions for this category
            category_questions = questions.filter(category=category)
            
            for q_idx, question in enumerate(category_questions, 1):
                question_responses = responses.filter(question=question)
                response_count = question_responses.count()
                
                # Question header
                question_text = f"{cat_idx}.{q_idx}. {question.question_text}"
                story.append(Paragraph(question_text, question_heading_style))
                
                # Question metadata
                meta_text = f"<i>Type: {question.get_question_type_display()} | Responses: {response_count}"
                if response_count > 0 and total_alumni > 0:
                    q_response_rate = round((response_count / total_alumni * 100), 1)
                    meta_text += f" | Response Rate: {q_response_rate}%"
                meta_text += "</i>"
                story.append(Paragraph(meta_text, styles['Normal']))
                story.append(Spacer(1, 0.1*inch))
                
                # Process based on question type
                if question.question_type == 'checkbox':
                    # Checkbox - HORIZONTAL BAR CHART
                    distribution = {}
                    for response in question_responses:
                        values = _extract_value_for_pdf(response.response_data)
                        if isinstance(values, list):
                            for val in values:
                                distribution[str(val)] = distribution.get(str(val), 0) + 1
                        elif values:
                            distribution[str(values)] = distribution.get(str(values), 0) + 1
                    
                    # Include all options even with 0 count
                    options = question.get_options_list()
                    for option in options:
                        if str(option) not in distribution:
                            distribution[str(option)] = 0
                    
                    if distribution:
                        sorted_dist = sorted(distribution.items(), key=lambda x: x[1], reverse=True)
                        
                        chart_height = max(150, len(sorted_dist) * 30)
                        drawing = Drawing(500, chart_height)
                        chart = HorizontalBarChart()
                        chart.x = 120
                        chart.y = 20
                        chart.height = chart_height - 40
                        chart.width = 350
                        
                        chart.data = [[item[1] for item in sorted_dist]]
                        chart.categoryAxis.categoryNames = [item[0] for item in sorted_dist]
                        
                        chart.bars[0].fillColor = colors.HexColor('#3b82f6')
                        chart.valueAxis.valueMin = 0
                        chart.categoryAxis.labels.fontSize = 8
                        chart.valueAxis.labels.fontSize = 8
                        
                        drawing.add(chart)
                        story.append(drawing)
                        story.append(Spacer(1, 0.1*inch))
                        
                        dist_data = [['Option', 'Count', 'Percentage']]
                        for option, count in sorted_dist:
                            percentage = round((count / response_count * 100), 1) if response_count > 0 else 0
                            dist_data.append([option, str(count), f"{percentage}%"])
                        
                        dist_table = Table(dist_data, colWidths=[2.5*inch, 1.5*inch, 1.7*inch])
                        dist_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#60a5fa')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
                        ]))
                        story.append(dist_table)
                
                elif question.question_type in ['radio', 'select']:
                    # Radio/Select - PIE CHART
                    distribution = {}
                    for response in question_responses:
                        value = _extract_value_for_pdf(response.response_data)
                        if value:
                            distribution[str(value)] = distribution.get(str(value), 0) + 1
                    
                    options = question.get_options_list()
                    for option in options:
                        if str(option) not in distribution:
                            distribution[str(option)] = 0
                    
                    if distribution and response_count > 0:
                        drawing = Drawing(450, 200)
                        pie = Pie()
                        pie.x = 100
                        pie.y = 20
                        pie.width = 150
                        pie.height = 150
                        
                        pie.data = list(distribution.values())
                        pie.labels = [f"{k}: {v}" for k, v in distribution.items()]
                        
                        color_palette = [
                            colors.HexColor('#a855f7'), colors.HexColor('#ec4899'),
                            colors.HexColor('#f97316'), colors.HexColor('#eab308'),
                            colors.HexColor('#84cc16'), colors.HexColor('#22c55e'),
                            colors.HexColor('#14b8a6'), colors.HexColor('#06b6d4'),
                            colors.HexColor('#3b82f6'), colors.HexColor('#6366f1'),
                        ]
                        for i, slice_color in enumerate(color_palette[:len(pie.data)]):
                            pie.slices[i].fillColor = slice_color
                        
                        drawing.add(pie)
                        story.append(drawing)
                        story.append(Spacer(1, 0.1*inch))
                        
                        dist_data = [['Option', 'Count', 'Percentage']]
                        for option, count in distribution.items():
                            percentage = round((count / response_count * 100), 1)
                            dist_data.append([option, str(count), f"{percentage}%"])
                        
                        dist_table = Table(dist_data, colWidths=[2.5*inch, 1.5*inch, 1.7*inch])
                        dist_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#60a5fa')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
                        ]))
                        story.append(dist_table)
                
                elif question.question_type == 'yes_no':
                    # Yes/No - PIE CHART with green/red colors
                    distribution = {'Yes': 0, 'No': 0}
                    for response in question_responses:
                        value = _extract_value_for_pdf(response.response_data)
                        if value in ['Yes', 'yes', True, 'true', '1', 1]:
                            distribution['Yes'] += 1
                        elif value in ['No', 'no', False, 'false', '0', 0]:
                            distribution['No'] += 1
                    
                    if response_count > 0:
                        drawing = Drawing(450, 200)
                        pie = Pie()
                        pie.x = 150
                        pie.y = 20
                        pie.width = 150
                        pie.height = 150
                        
                        pie.data = [distribution['Yes'], distribution['No']]
                        pie.labels = [f"Yes: {distribution['Yes']}", f"No: {distribution['No']}"]
                        
                        pie.slices[0].fillColor = colors.HexColor('#22c55e')
                        pie.slices[1].fillColor = colors.HexColor('#ef4444')
                        
                        drawing.add(pie)
                        story.append(drawing)
                        story.append(Spacer(1, 0.1*inch))
                        
                        dist_data = [
                            ['Response', 'Count', 'Percentage'],
                            ['Yes', str(distribution['Yes']), f"{round(distribution['Yes']/response_count*100, 1)}%"],
                            ['No', str(distribution['No']), f"{round(distribution['No']/response_count*100, 1)}%"]
                        ]
                        
                        dist_table = Table(dist_data, colWidths=[2.5*inch, 1.5*inch, 1.7*inch])
                        dist_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#60a5fa')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
                        ]))
                        story.append(dist_table)
                
                elif question.question_type == 'rating':
                    # Rating scale - BAR CHART with average
                    values = []
                    distribution = {}
                    
                    for response in question_responses:
                        value = _extract_value_for_pdf(response.response_data)
                        if value is not None:
                            try:
                                numeric_value = float(value)
                                values.append(numeric_value)
                                rating_key = int(numeric_value)
                                distribution[rating_key] = distribution.get(rating_key, 0) + 1
                            except (ValueError, TypeError):
                                pass
                    
                    if values:
                        avg = round(sum(values) / len(values), 2)
                        min_val = question.min_value or 1
                        max_val = question.max_value or 5
                        
                        avg_data = [[f'Average Rating: {avg} out of {max_val}']]
                        avg_table = Table(avg_data, colWidths=[5.7*inch])
                        avg_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fef3c7')),
                            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#92400e')),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 14),
                            ('TOPPADDING', (0, 0), (-1, -1), 12),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#fbbf24')),
                        ]))
                        story.append(avg_table)
                        story.append(Spacer(1, 0.15*inch))
                        
                        for rating in range(min_val, max_val + 1):
                            if rating not in distribution:
                                distribution[rating] = 0
                        
                        drawing = Drawing(500, 200)
                        chart = VerticalBarChart()
                        chart.x = 50
                        chart.y = 20
                        chart.height = 150
                        chart.width = 400
                        
                        sorted_ratings = sorted(distribution.items())
                        chart.data = [[item[1] for item in sorted_ratings]]
                        chart.categoryAxis.categoryNames = [f"{item[0]} {'star' if item[0] == 1 else 'stars'}" for item in sorted_ratings]
                        
                        chart.bars[0].fillColor = colors.HexColor('#fbbf24')
                        chart.valueAxis.valueMin = 0
                        chart.categoryAxis.labels.angle = 0
                        chart.categoryAxis.labels.fontSize = 8
                        chart.valueAxis.labels.fontSize = 8
                        
                        drawing.add(chart)
                        story.append(drawing)
                        story.append(Spacer(1, 0.1*inch))
                        
                        dist_data = [['Rating', 'Count', 'Percentage']]
                        for rating, count in sorted_ratings:
                            percentage = round((count / len(values) * 100), 1)
                            dist_data.append([f"{rating} {'star' if rating == 1 else 'stars'}", str(count), f"{percentage}%"])
                        dist_data.append(['Average', str(avg), ''])
                        
                        dist_table = Table(dist_data, colWidths=[2.5*inch, 1.5*inch, 1.7*inch])
                        dist_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#60a5fa')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef3c7')),
                            ('FONTSIZE', (0, 0), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#eff6ff')]),
                        ]))
                        story.append(dist_table)
                
                elif question.question_type == 'number':
                    # Number questions - BAR CHART with statistics
                    values = []
                    distribution = {}
                    
                    for response in question_responses:
                        value = _extract_value_for_pdf(response.response_data)
                        if value is not None:
                            try:
                                num_value = float(value)
                                values.append(num_value)
                                distribution[num_value] = distribution.get(num_value, 0) + 1
                            except (ValueError, TypeError):
                                pass
                    
                    if values:
                        avg = round(sum(values) / len(values), 2)
                        min_val = round(min(values), 2)
                        max_val = round(max(values), 2)
                        
                        stats_data = [
                            ['Average', 'Minimum', 'Maximum'],
                            [str(avg), str(min_val), str(max_val)]
                        ]
                        stats_table = Table(stats_data, colWidths=[1.9*inch, 1.9*inch, 1.9*inch])
                        stats_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14b8a6')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ccfbf1')),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 11),
                            ('TOPPADDING', (0, 0), (-1, -1), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#14b8a6')),
                        ]))
                        story.append(stats_table)
                        story.append(Spacer(1, 0.15*inch))
                        
                        if len(distribution) <= 15:
                            drawing = Drawing(500, 200)
                            chart = VerticalBarChart()
                            chart.x = 50
                            chart.y = 20
                            chart.height = 150
                            chart.width = 400
                            
                            sorted_dist = sorted(distribution.items())
                            chart.data = [[item[1] for item in sorted_dist]]
                            chart.categoryAxis.categoryNames = [str(item[0]) for item in sorted_dist]
                            
                            chart.bars[0].fillColor = colors.HexColor('#14b8a6')
                            chart.valueAxis.valueMin = 0
                            chart.categoryAxis.labels.angle = 45
                            chart.categoryAxis.labels.fontSize = 7
                            chart.valueAxis.labels.fontSize = 8
                            
                            drawing.add(chart)
                            story.append(drawing)
                            story.append(Spacer(1, 0.1*inch))
                            
                            dist_data = [['Value', 'Count', 'Percentage']]
                            for value, count in sorted_dist:
                                percentage = round((count / len(values) * 100), 1)
                                dist_data.append([str(value), str(count), f"{percentage}%"])
                            
                            dist_table = Table(dist_data, colWidths=[2*inch, 1.5*inch, 2.2*inch])
                            dist_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#60a5fa')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('FONTSIZE', (0, 0), (-1, -1), 9),
                                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
                            ]))
                            story.append(dist_table)
                
                elif question.question_type == 'year':
                    # Year questions - BAR CHART
                    distribution = {}
                    
                    for response in question_responses:
                        value = _extract_value_for_pdf(response.response_data)
                        if value:
                            distribution[str(value)] = distribution.get(str(value), 0) + 1
                    
                    if distribution:
                        sorted_years = sorted(distribution.items(), key=lambda x: x[0], reverse=True)
                        
                        drawing = Drawing(500, 200)
                        chart = VerticalBarChart()
                        chart.x = 50
                        chart.y = 20
                        chart.height = 150
                        chart.width = 400
                        
                        chart.data = [[item[1] for item in sorted_years]]
                        chart.categoryAxis.categoryNames = [item[0] for item in sorted_years]
                        
                        chart.bars[0].fillColor = colors.HexColor('#10b981')
                        chart.valueAxis.valueMin = 0
                        chart.categoryAxis.labels.angle = 45
                        chart.categoryAxis.labels.fontSize = 8
                        chart.valueAxis.labels.fontSize = 8
                        
                        drawing.add(chart)
                        story.append(drawing)
                        story.append(Spacer(1, 0.1*inch))
                        
                        dist_data = [['Year', 'Count', 'Percentage']]
                        for year, count in sorted_years:
                            percentage = round((count / response_count * 100), 1)
                            dist_data.append([year, str(count), f"{percentage}%"])
                        
                        dist_table = Table(dist_data, colWidths=[2.5*inch, 1.5*inch, 1.7*inch])
                        dist_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#60a5fa')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
                        ]))
                        story.append(dist_table)
                
                elif question.question_type in ['text', 'textarea', 'email']:
                    # Text responses - show response count only (privacy)
                    count_data = [[f'{response_count} text responses received']]
                    count_table = Table(count_data, colWidths=[5.7*inch])
                    count_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, -1), 11),
                        ('TOPPADDING', (0, 0), (-1, -1), 15),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
                        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#60a5fa')),
                    ]))
                    story.append(count_table)
                    story.append(Paragraph("<i>Individual responses are protected for privacy</i>", styles['Normal']))
                
                story.append(Spacer(1, 0.25*inch))
            
            # Add page break after each category except the last
            if cat_idx < len(categories):
                story.append(PageBreak())
        
        # Build PDF
        doc.build(story)
        
        # Return response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # Use form title in filename
        safe_form_name = ''.join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in form_title)
        filename = f'{safe_form_name}_Complete_Analytics_{timestamp}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError:
        return Response({
            'error': 'ReportLab library not installed. Run: pip install reportlab'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': f'PDF export failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# CACHE MANAGEMENT
# =============================================================================

@api_view(['POST'])
@permission_classes([IsSuperAdminOnly])
def clear_survey_cache_view(request):
    """
    Clear all survey-related cache.
    Super admin only utility endpoint.
    NOTE: Registration survey no longer uses caching.
    """
    return Response({
        'message': 'No cache clearing needed - registration survey fetches live data',
        'cleared_at': timezone.now().isoformat()
    })
