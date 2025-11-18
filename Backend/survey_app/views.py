from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db.models import Count, Q
from django.utils import timezone
from django.core.cache import cache
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator

from .models import SurveyCategory, SurveyQuestion, SurveyResponse, SurveyTemplate
from .serializers import (
    SurveyCategorySerializer, SurveyQuestionSerializer, SurveyQuestionListSerializer,
    SurveyResponseSerializer, SurveyResponseSubmissionSerializer,
    ActiveSurveyQuestionsSerializer, SurveyAnalyticsSerializer,
    SurveyTemplateSerializer
)
from .permissions import IsSurveyAdmin, IsSuperAdminOnly, CanRespondToSurveys, IsSurveyOwnerOrAdmin

from rest_framework import generics
from django.shortcuts import get_object_or_404


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

# Note: Registration survey no longer uses caching for immediate updates
    
    # Clear user-specific active survey caches (pattern-based)
    # Note: In production, you might want to use cache.delete_pattern if available
    # For now, we'll clear the main caches which will force refresh for all users


# =============================================================================
# ADMIN VIEWS - Survey Management (Super Admin & Admin Only)
# =============================================================================

class SurveyCategoryListCreateView(generics.ListCreateAPIView):
    """
    List all survey categories or create a new one.
    Only admins can access this endpoint.
    """
    serializer_class = SurveyCategorySerializer
    permission_classes = [IsSurveyAdmin]
    
    def get_queryset(self):
        return SurveyCategory.objects.all().order_by('order', 'name')
    
    def perform_create(self, serializer):
        """Save category with audit information"""
        super().perform_create(serializer)
        # No cache clearing needed - registration endpoint doesn't use cache


class SurveyCategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update or delete a survey category.
    Only super admins can delete categories.
    """
    serializer_class = SurveyCategorySerializer
    queryset = SurveyCategory.objects.all()
    
    def get_permissions(self):
        if self.request.method == 'DELETE':
            return [IsSuperAdminOnly()]
        return [IsSurveyAdmin()]
    
    def perform_update(self, serializer):
        """Update category"""
        super().perform_update(serializer)
        # No cache clearing needed - registration endpoint doesn't use cache
    
    def perform_destroy(self, instance):
        """Delete category"""
        super().perform_destroy(instance)
        # No cache clearing needed - registration endpoint doesn't use cache


class SurveyQuestionListCreateView(generics.ListCreateAPIView):
    """
    List all survey questions or create a new one.
    Supports filtering by category and active status.
    """
    serializer_class = SurveyQuestionSerializer
    permission_classes = [IsSurveyAdmin]
    
    def get_queryset(self):
        queryset = SurveyQuestion.objects.select_related('category', 'created_by')
        
        # Filter by category
        category_id = self.request.query_params.get('category')
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('category__order', 'order', 'question_text')
    
    def perform_create(self, serializer):
        """Save question with audit information"""
        super().perform_create(serializer)
        # No cache clearing needed - registration endpoint doesn't use cache


class SurveyQuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update or delete a survey question.
    Only super admins can delete questions.
    """
    serializer_class = SurveyQuestionSerializer
    queryset = SurveyQuestion.objects.all()
    
    def get_permissions(self):
        if self.request.method == 'DELETE':
            return [IsSuperAdminOnly()]
        return [IsSurveyAdmin()]
    
    def perform_update(self, serializer):
        """Update question"""
        super().perform_update(serializer)
        # No cache clearing needed - registration endpoint doesn't use cache
    
    def perform_destroy(self, instance):
        """Delete question"""
        super().perform_destroy(instance)
        # No cache clearing needed - registration endpoint doesn't use cache


# -----------------------------------------------------------------------------
# Form (SurveyTemplate) Management - Treat SurveyTemplate as top-level Form
# -----------------------------------------------------------------------------


class SurveyFormListCreateView(generics.ListCreateAPIView):
    """List all forms (templates) or create a new form.
    Uses SurveyTemplate as the Form model.
    """
    serializer_class = SurveyTemplateSerializer
    permission_classes = [IsSurveyAdmin]

    def get_queryset(self):
        return SurveyTemplate.objects.all().order_by('name')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class SurveyFormDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete a form/template."""
    serializer_class = SurveyTemplateSerializer
    queryset = SurveyTemplate.objects.all()

    def get_permissions(self):
        if self.request.method == 'DELETE':
            return [IsSuperAdminOnly()]
        return [IsSurveyAdmin()]

    def retrieve(self, request, *args, **kwargs):
        # Return nested form -> categories -> questions structure
        instance = self.get_object()
        form_data = SurveyTemplateSerializer(instance).data

        # Attach categories with their questions
        categories = instance.categories.all().order_by('order', 'name')
        categories_data = []
        for cat in categories:
            questions = SurveyQuestion.objects.filter(category=cat).order_by('order', 'question_text')
            qdata = SurveyQuestionListSerializer(questions, many=True).data
            categories_data.append({
                'category': SurveyCategorySerializer(cat).data,
                'questions': qdata
            })

        form_data['sections'] = categories_data
        return Response(form_data)

    def perform_update(self, serializer):
        # Allow updating template fields and categories (category_ids handled in serializer)
        serializer.save()


class SurveyFormPublishView(APIView):
    """Toggle publish or accepting responses for a form."""
    permission_classes = [IsSurveyAdmin]

    def post(self, request, pk):
        form = get_object_or_404(SurveyTemplate, pk=pk)
        is_published = request.data.get('is_published')
        accepting = request.data.get('accepting_responses')

        if is_published is not None:
            form.is_published = bool(is_published)
        if accepting is not None:
            form.accepting_responses = bool(accepting)
        form.save()

        return Response({'message': 'Form updated', 'is_published': form.is_published, 'accepting_responses': form.accepting_responses})


class SurveyResponseAnalyticsView(APIView):
    """
    Provide analytics data for survey responses.
    Admin-only endpoint for viewing survey statistics.
    """
    permission_classes = [IsSurveyAdmin]
    
    def get(self, request):
        cache_key = 'survey_analytics_data'
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
        
        # Calculate analytics
        total_questions = SurveyQuestion.objects.filter(is_active=True).count()
        total_responses = SurveyResponse.objects.count()
        total_users_responded = SurveyResponse.objects.values('user').distinct().count()
        
        # Calculate completion rate
        from django.contrib.auth import get_user_model
        User = get_user_model()
        total_alumni = User.objects.filter(user_type=3, is_approved=True).count()
        completion_rate = (total_users_responded / total_alumni * 100) if total_alumni > 0 else 0
        
        # Category statistics
        category_stats = []
        for category in SurveyCategory.objects.filter(is_active=True):
            questions_count = category.questions.filter(is_active=True).count()
            responses_count = SurveyResponse.objects.filter(
                question__category=category
            ).count()
            category_stats.append({
                'name': category.name,
                'questions_count': questions_count,
                'responses_count': responses_count,
                'response_rate': (responses_count / (questions_count * total_alumni) * 100) if questions_count > 0 and total_alumni > 0 else 0
            })
        
        # Question statistics (top 10 most answered)
        question_stats = []
        top_questions = SurveyQuestion.objects.filter(is_active=True).annotate(
            responses_total=Count('responses')
        ).order_by('-responses_total')[:10]
        
        for question in top_questions:
            question_stats.append({
                'id': question.id,
                'question_text': question.question_text[:50] + '...' if len(question.question_text) > 50 else question.question_text,
                'category': question.category.name,
                'response_count': question.responses_total,
                'response_rate': (question.responses_total / total_alumni * 100) if total_alumni > 0 else 0
            })
        
        analytics_data = {
            'total_questions': total_questions,
            'total_responses': total_responses,
            'total_users_responded': total_users_responded,
            'completion_rate': round(completion_rate, 2),
            'category_stats': category_stats,
            'question_stats': question_stats,
            'generated_at': timezone.now().isoformat()
        }
        
        # Cache for 1 hour
        cache.set(cache_key, analytics_data, 3600)
        
        return Response(analytics_data)


class SurveyResponsesView(generics.ListAPIView):
    """
    List all survey responses for admin review.
    Supports filtering by user, question, and category.
    """
    serializer_class = SurveyResponseSerializer
    permission_classes = [IsSurveyAdmin]
    
    def get_queryset(self):
        queryset = SurveyResponse.objects.select_related(
            'user', 'question', 'question__category'
        )
        
        # Filter by user
        user_id = self.request.query_params.get('user_id')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by question
        question_id = self.request.query_params.get('question_id')
        if question_id:
            queryset = queryset.filter(question_id=question_id)
        
        # Filter by category
        category_id = self.request.query_params.get('category_id')
        if category_id:
            queryset = queryset.filter(question__category_id=category_id)
        
        return queryset.order_by('-submitted_at')


class CategoryAnalyticsView(APIView):
    """
    Get detailed analytics for a specific category.
    Returns per-question aggregated statistics ready for charting.
    """
    permission_classes = [IsSurveyAdmin]
    
    def get(self, request):
        category_id = request.query_params.get('category_id')
        
        if not category_id:
            return Response(
                {'error': 'category_id parameter is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            category = SurveyCategory.objects.get(id=category_id)
        except SurveyCategory.DoesNotExist:
            return Response(
                {'error': 'Category not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check cache first
        cache_key = f'category_analytics_{category_id}'
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)
        
        # Get all questions for this category
        questions = SurveyQuestion.objects.filter(
            category=category,
            is_active=True
        ).order_by('order', 'question_text')
        
        # Get all responses for this category
        responses = SurveyResponse.objects.filter(
            question__category=category
        ).select_related('user', 'question')
        
        # Get total potential respondents (alumni)
        from django.contrib.auth import get_user_model
        User = get_user_model()
        total_alumni = User.objects.filter(user_type=3, is_approved=True).count()
        
        # Calculate per-question analytics
        question_analytics = []
        
        for question in questions:
            question_responses = responses.filter(question=question)
            response_count = question_responses.count()
            
            analytics_item = {
                'question_id': question.id,
                'question_text': question.question_text,
                'question_type': question.question_type,
                'help_text': question.help_text,
                'is_required': question.is_required,
                'response_count': response_count,
                'response_rate': round((response_count / total_alumni * 100) if total_alumni > 0 else 0, 2),
                'order': question.order
            }
            
            # Type-specific analytics
            if question.question_type in ['radio', 'select']:
                # Single choice - calculate distribution
                distribution = {}
                for response in question_responses:
                    value = self._extract_value(response.response_data)
                    if value:
                        distribution[str(value)] = distribution.get(str(value), 0) + 1
                
                analytics_item['distribution'] = distribution
                analytics_item['options'] = question.get_options_list()
                
            elif question.question_type == 'checkbox':
                # Multiple choice - calculate distribution (multi-select)
                distribution = {}
                for response in question_responses:
                    values = self._extract_value(response.response_data)
                    if isinstance(values, list):
                        for value in values:
                            distribution[str(value)] = distribution.get(str(value), 0) + 1
                    elif values:
                        distribution[str(values)] = distribution.get(str(values), 0) + 1
                
                analytics_item['distribution'] = distribution
                analytics_item['options'] = question.get_options_list()
                
            elif question.question_type == 'rating':
                # Rating scale - calculate distribution and average
                values = []
                distribution = {}
                
                for response in question_responses:
                    value = self._extract_value(response.response_data)
                    if value is not None:
                        try:
                            numeric_value = float(value)
                            values.append(numeric_value)
                            distribution[str(int(numeric_value))] = distribution.get(str(int(numeric_value)), 0) + 1
                        except (ValueError, TypeError):
                            pass
                
                analytics_item['distribution'] = distribution
                analytics_item['average'] = round(sum(values) / len(values), 2) if values else 0
                analytics_item['min_value'] = question.min_value or 1
                analytics_item['max_value'] = question.max_value or 5
                
            elif question.question_type == 'yes_no':
                # Yes/No - calculate distribution
                distribution = {'Yes': 0, 'No': 0}
                for response in question_responses:
                    value = self._extract_value(response.response_data)
                    if value in ['Yes', 'yes', True, 'true', '1', 1]:
                        distribution['Yes'] += 1
                    elif value in ['No', 'no', False, 'false', '0', 0]:
                        distribution['No'] += 1
                
                analytics_item['distribution'] = distribution
                
            elif question.question_type in ['text', 'textarea', 'email']:
                # Text responses - provide samples
                sample_responses = []
                for response in question_responses[:10]:  # Limit to 10 samples
                    value = self._extract_value(response.response_data)
                    if value:
                        sample_responses.append({
                            'value': str(value)[:200],  # Truncate long text
                            'submitted_at': response.submitted_at.isoformat()
                        })
                
                analytics_item['sample_responses'] = sample_responses
                
            elif question.question_type == 'number':
                # Number - calculate statistics and create histogram buckets
                values = []
                for response in question_responses:
                    value = self._extract_value(response.response_data)
                    if value is not None:
                        try:
                            values.append(float(value))
                        except (ValueError, TypeError):
                            pass
                
                if values:
                    analytics_item['average'] = round(sum(values) / len(values), 2)
                    analytics_item['min'] = min(values)
                    analytics_item['max'] = max(values)
                    
                    # Create histogram buckets for bar chart
                    min_val = min(values)
                    max_val = max(values)
                    range_val = max_val - min_val
                    
                    if range_val == 0:
                        # All same value
                        analytics_item['distribution'] = {str(int(min_val)): len(values)}
                    else:
                        # Determine number of buckets (5-10 depending on range)
                        num_buckets = min(10, max(5, int(range_val / 5) + 1))
                        bucket_width = range_val / num_buckets
                        
                        # Create buckets
                        distribution = {}
                        for i in range(num_buckets):
                            bucket_start = min_val + (i * bucket_width)
                            bucket_end = bucket_start + bucket_width
                            
                            if i == num_buckets - 1:
                                # Last bucket includes max value
                                bucket_label = f"{int(bucket_start)}-{int(bucket_end)}"
                                count = sum(1 for v in values if bucket_start <= v <= bucket_end)
                            else:
                                bucket_label = f"{int(bucket_start)}-{int(bucket_end-1)}"
                                count = sum(1 for v in values if bucket_start <= v < bucket_end)
                            
                            if count > 0:  # Only include non-empty buckets
                                distribution[bucket_label] = count
                        
                        analytics_item['distribution'] = distribution
                else:
                    analytics_item['average'] = 0
                    analytics_item['min'] = 0
                    analytics_item['max'] = 0
                    analytics_item['distribution'] = {}
                
            elif question.question_type == 'date':
                # Date - provide distribution by year/month
                dates = []
                for response in question_responses:
                    value = self._extract_value(response.response_data)
                    if value:
                        dates.append(str(value))
                
                analytics_item['sample_dates'] = dates[:10]
            
            question_analytics.append(analytics_item)
        
        # Category summary
        category_data = {
            'category': {
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'total_questions': questions.count(),
                'active_questions': questions.count()
            },
            'summary': {
                'total_responses': responses.count(),
                'unique_respondents': responses.values('user').distinct().count(),
                'total_potential_respondents': total_alumni,
                'response_rate': round(
                    (responses.values('user').distinct().count() / total_alumni * 100) 
                    if total_alumni > 0 else 0, 
                    2
                )
            },
            'questions': question_analytics,
            'generated_at': timezone.now().isoformat()
        }
        
        # Cache for 30 minutes
        cache.set(cache_key, category_data, 1800)
        
        return Response(category_data)
    
    def _extract_value(self, response_data):
        """Extract the actual value from response_data JSON"""
        if not response_data:
            return None
        
        if isinstance(response_data, dict):
            # Try common keys
            for key in ['value', 'answer', 'text', 'rating', 'selected_options']:
                if key in response_data:
                    return response_data[key]
            return response_data
        
        return response_data


# =============================================================================
# ALUMNI VIEWS - Survey Taking (Alumni & Authenticated Users)
# =============================================================================

class ActiveSurveyQuestionsView(APIView):
    """
    Get all active survey questions organized by categories.
    Alumni use this to see what surveys they can fill out.
    """
    permission_classes = [CanRespondToSurveys]
    
    def get(self, request):
        cache_key = f'active_survey_questions_user_{request.user.id}'
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
        
        # Get active categories with their questions
        categories = SurveyCategory.objects.filter(is_active=True).prefetch_related(
            'questions'
        ).order_by('order', 'name')
        
        survey_data = []
        for category in categories:
            active_questions = category.questions.filter(is_active=True).order_by('order', 'question_text')
            
            if active_questions.exists():
                questions_data = ActiveSurveyQuestionsSerializer(
                    active_questions, 
                    many=True, 
                    context={'request': request}
                ).data
                
                survey_data.append({
                    'category': SurveyCategorySerializer(category).data,
                    'questions': questions_data
                })
        
        # Cache for 30 minutes
        cache.set(cache_key, survey_data, 1800)
        
        return Response(survey_data)


class SurveyResponseSubmitView(APIView):
    """
    Submit survey responses (single or multiple).
    Alumni use this to submit their survey answers.
    """
    permission_classes = [CanRespondToSurveys]
    
    def post(self, request):
        # Debug logging
        print(f"🚀 Survey submission received from user: {request.user}")
        print(f"📋 Request data: {request.data}")
        print(f"📊 Data type: {type(request.data)}")
        
        # Check if it's bulk submission
        if 'responses' in request.data:
            print("📦 Processing as bulk submission")
            serializer = SurveyResponseSubmissionSerializer(
                data=request.data,
                context={'request': request, 'ip_address': self.get_client_ip(request)}
            )
        else:
            print("📝 Processing as single response")
            # Single response submission
            serializer = SurveyResponseSerializer(
                data=request.data,
                context={'request': request}
            )
        
        if serializer.is_valid():
            print("✅ Serializer validation passed")
            response = serializer.save()
            print(f"✅ Response saved successfully: {response}")
            
            # Clear user's cache
            cache_key = f'active_survey_questions_user_{request.user.id}'
            cache.delete(cache_key)
            cache.delete('survey_analytics_data')  # Clear analytics cache
            
            return Response(
                {'message': 'Survey response(s) submitted successfully'},
                status=status.HTTP_201_CREATED
            )
        
        print(f"❌ Serializer validation failed: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            return x_forwarded_for.split(',')[0]
        return request.META.get('REMOTE_ADDR')


class UserSurveyResponsesView(generics.ListAPIView):
    """
    Get current user's survey responses.
    Alumni can view their own submitted responses.
    """
    serializer_class = SurveyResponseSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return SurveyResponse.objects.filter(
            user=self.request.user
        ).select_related('question', 'question__category').order_by('-submitted_at')


class SurveyProgressView(APIView):
    """
    Get user's survey completion progress.
    Shows how many questions they've answered vs total questions.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        # Get total active questions
        total_questions = SurveyQuestion.objects.filter(is_active=True).count()
        
        # Get user's responses count
        answered_questions = SurveyResponse.objects.filter(user=user).count()
        
        # Calculate progress percentage
        progress_percentage = (answered_questions / total_questions * 100) if total_questions > 0 else 0
        
        # Get category-wise progress
        category_progress = []
        for category in SurveyCategory.objects.filter(is_active=True):
            category_questions = category.questions.filter(is_active=True).count()
            category_answered = SurveyResponse.objects.filter(
                user=user, 
                question__category=category
            ).count()
            category_progress.append({
                'category_name': category.name,
                'total_questions': category_questions,
                'answered_questions': category_answered,
                'progress_percentage': (category_answered / category_questions * 100) if category_questions > 0 else 0
            })
        
        return Response({
            'total_questions': total_questions,
            'answered_questions': answered_questions,
            'progress_percentage': round(progress_percentage, 2),
            'category_progress': category_progress,
            'is_complete': answered_questions >= total_questions
        })


# =============================================================================
# UTILITY VIEWS
# =============================================================================

@api_view(['POST'])
@permission_classes([IsSurveyAdmin])
def survey_export_view(request):
    """
    🚀 DYNAMIC EXPORT - Automatically includes ALL survey questions, user fields, and profile data.
    Adapts to any changes in questions or user models without code changes.
    Includes COMPLETE alumni directory information for comprehensive tracing.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.apps import apps
    import io
    import json
    from datetime import datetime
    
    try:
        # Get parameters from request body
        export_format = request.data.get('format', 'xlsx')
        category_id = request.data.get('category_id')
        date_from = request.data.get('date_from')
        date_to = request.data.get('date_to')
        
        print(f"🔍 Starting DYNAMIC export with filters: category={category_id}, date_from={date_from}, date_to={date_to}")
        
        # ===== STEP 1: Get ALL survey questions dynamically (not hardcoded) =====
        all_questions = SurveyQuestion.objects.all().select_related('category').order_by('category__name', 'order')
        if category_id:
            all_questions = all_questions.filter(category_id=category_id)
        
        print(f"📊 Found {all_questions.count()} questions to include")
        
        # ===== STEP 2: Get ALL survey responses dynamically =====
        all_responses = SurveyResponse.objects.select_related('user', 'question', 'question__category')
        
        # Apply date filters
        if date_from:
            all_responses = all_responses.filter(submitted_at__gte=date_from)
        if date_to:
            all_responses = all_responses.filter(submitted_at__lte=date_to)
        
        # Apply category filter
        if category_id:
            all_responses = all_responses.filter(question__category_id=category_id)
        
        print(f"📝 Found {all_responses.count()} responses to include")
        
        # ===== STEP 3: Get ALL users who have ANY responses =====
        # Ensure User model is resolved (use get_user_model to support custom user model)
        from django.contrib.auth import get_user_model
        User = get_user_model()

        user_ids = all_responses.values_list('user_id', flat=True).distinct()
        # auth_app.Address uses related_name='normalized_addresses'
        users = User.objects.filter(id__in=user_ids).select_related('profile').prefetch_related('normalized_addresses')
        
        print(f"👥 Found {users.count()} users with responses")
        
        if not users.exists():
            return Response({'error': 'No survey responses found for the specified criteria.'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # ===== STEP 4: DYNAMICALLY get ALL fields from ALL models =====
        
        # DYNAMICALLY get ALL fields from User model
        user_model_fields = []
        user_excluded = ['password', 'last_login', 'is_superuser', 'is_staff', 'groups', 'user_permissions']
        for field in User._meta.fields:
            if field.name not in user_excluded:
                user_model_fields.append({
                    'name': field.name,
                    'verbose_name': field.verbose_name or field.name.replace('_', ' ').title()
                })
        
        # DYNAMICALLY get ALL fields from Profile model
        profile_model_fields = []
        try:
            Profile = apps.get_model('auth_app', 'Profile')
            for field in Profile._meta.fields:
                if field.name not in ['id', 'user']:
                    profile_model_fields.append({
                        'name': field.name,
                        'verbose_name': field.verbose_name or field.name.replace('_', ' ').title()
                    })
        except:
            print("⚠️ Profile model not found")
        
        # DYNAMICALLY get ALL fields from Address model
        address_model_fields = []
        try:
            Address = apps.get_model('auth_app', 'Address')
            for field in Address._meta.fields:
                if field.name not in ['id', 'user']:
                    address_model_fields.append({
                        'name': field.name,
                        'verbose_name': field.verbose_name or field.name.replace('_', ' ').title()
                    })
        except:
            print("⚠️ Address model not found")
        
        # ===== STEP 5: Create workbook =====
        wb = Workbook()
        
        # ===== SHEET 1: COMPLETE ALUMNI DATA =====
        ws_main = wb.active
        ws_main.title = "Complete Alumni Data"
        
        # Build DYNAMIC headers
        headers = []
        header_mapping = {}  # Track what each column represents
        col_index = 0
        
        # Add ALL User model fields
        for field_info in user_model_fields:
            headers.append(f"USER: {field_info['verbose_name']}")
            header_mapping[col_index] = ('user', field_info['name'])
            col_index += 1
        
        # Add ALL Profile model fields
        for field_info in profile_model_fields:
            headers.append(f"PROFILE: {field_info['verbose_name']}")
            header_mapping[col_index] = ('profile', field_info['name'])
            col_index += 1
        
        # Add ALL Address model fields (Present Address)
        for field_info in address_model_fields:
            headers.append(f"PRESENT ADDRESS: {field_info['verbose_name']}")
            header_mapping[col_index] = ('address_present', field_info['name'])
            col_index += 1
        
        # Add ALL Address model fields (Permanent Address)
        for field_info in address_model_fields:
            headers.append(f"PERMANENT ADDRESS: {field_info['verbose_name']}")
            header_mapping[col_index] = ('address_permanent', field_info['name'])
            col_index += 1
        
        # Add ALL Survey Questions DYNAMICALLY
        question_columns = {}
        for question in all_questions:
            question_header = f"SURVEY: [{question.category.name}] {question.question_text}"
            if len(question_header) > 100:
                question_header = question_header[:97] + "..."
            headers.append(question_header)
            header_mapping[col_index] = ('question', question.id)
            question_columns[question.id] = col_index
            col_index += 1
        
        print(f"📋 Created {len(headers)} dynamic columns")
        
        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            
            # Different colors for different sections
            if header.startswith("USER:"):
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            elif header.startswith("PROFILE:"):
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            elif "ADDRESS:" in header:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            elif header.startswith("SURVEY:"):
                cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
        
        # ===== STEP 6: Get ALL responses organized by user and question =====
        user_responses = {}
        for response in all_responses:
            if response.user_id not in user_responses:
                user_responses[response.user_id] = {}
            user_responses[response.user_id][response.question_id] = response.response_data
        
        # ===== STEP 7: Write data for each user DYNAMICALLY =====
        for row, user in enumerate(users, 2):
            for col_idx, (data_type, field_name) in header_mapping.items():
                col = col_idx + 1
                cell_value = ""
                
                try:
                    if data_type == 'user':
                        # Get value from User model
                        cell_value = getattr(user, field_name, '')
                        
                    elif data_type == 'profile':
                        # Get value from Profile model
                        if hasattr(user, 'profile') and user.profile:
                            cell_value = getattr(user.profile, field_name, '')
                        
                    elif data_type == 'address_present':
                        # Get value from Present Address (Address model related_name: normalized_addresses)
                        present_address = user.normalized_addresses.filter(address_category='present').first()
                        if present_address:
                            cell_value = getattr(present_address, field_name, '')
                        
                    elif data_type == 'address_permanent':
                        # Get value from Permanent Address (Address model related_name: normalized_addresses)
                        permanent_address = user.normalized_addresses.filter(address_category='permanent').first()
                        if permanent_address:
                            cell_value = getattr(permanent_address, field_name, '')
                        
                    elif data_type == 'question':
                        # Get survey response
                        question_id = field_name
                        if user.id in user_responses and question_id in user_responses[user.id]:
                            response_data = user_responses[user.id][question_id]
                            
                            # DYNAMICALLY handle ALL response formats
                            if isinstance(response_data, dict):
                                if 'value' in response_data:
                                    cell_value = response_data['value']
                                elif 'selected_options' in response_data:
                                    cell_value = ', '.join(map(str, response_data['selected_options']))
                                elif 'rating' in response_data:
                                    cell_value = response_data['rating']
                                elif 'text' in response_data:
                                    cell_value = response_data['text']
                                elif 'answer' in response_data:
                                    cell_value = response_data['answer']
                                else:
                                    cell_value = json.dumps(response_data) if response_data else ''
                            elif isinstance(response_data, list):
                                cell_value = ', '.join(map(str, response_data))
                            else:
                                cell_value = str(response_data) if response_data else ''
                    
                    # Convert to string and handle None values
                    if cell_value is None:
                        cell_value = ''
                    elif isinstance(cell_value, (list, dict)):
                        cell_value = json.dumps(cell_value)
                    else:
                        cell_value = str(cell_value)
                    
                except Exception as e:
                    cell_value = f"Error: {str(e)}"
                
                # Write cell value
                ws_main.cell(row=row, column=col, value=cell_value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws_main.column_dimensions[column_letter].width = 20
        
        # ===== SHEET 2: QUESTIONS REFERENCE =====
        ws_questions = wb.create_sheet("Questions Reference")
        q_headers = ['Category', 'Question ID', 'Question Text', 'Question Type', 'Options/Scale', 'Is Required', 'Order', 'Response Count']
        
        # Write question headers
        for col, header in enumerate(q_headers, 1):
            cell = ws_questions.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Write question data
        for row, question in enumerate(all_questions, 2):
            response_count = all_responses.filter(question=question).count()
            
            # Get options/scale info
            options_info = ""
            if hasattr(question, 'options') and question.options:
                try:
                    options_data = json.loads(question.options) if isinstance(question.options, str) else question.options
                    if isinstance(options_data, list):
                        options_info = ', '.join(options_data)
                    elif isinstance(options_data, dict):
                        options_info = json.dumps(options_data)
                except:
                    options_info = str(question.options)
            
            question_data = [
                question.category.name,
                question.id,
                question.question_text,
                question.question_type,
                options_info,
                'Yes' if question.is_required else 'No',
                question.order,
                response_count
            ]
            
            for col, data in enumerate(question_data, 1):
                ws_questions.cell(row=row, column=col, value=str(data))
        
        # ===== SHEET 3: EXPORT SUMMARY =====
        ws_summary = wb.create_sheet("Export Summary")
        summary_data = [
            ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Users Exported', users.count()],
            ['Total Questions Included', all_questions.count()],
            ['Total Responses Included', all_responses.count()],
            ['Category Filter', category_id or 'All Categories'],
            ['Date From Filter', date_from or 'No Limit'],
            ['Date To Filter', date_to or 'No Limit'],
            ['User Model Fields', len(user_model_fields)],
            ['Profile Model Fields', len(profile_model_fields)],
            ['Address Model Fields', len(address_model_fields) * 2],  # Present + Permanent
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=str(value))
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'complete_alumni_data_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print(f"✅ DYNAMIC export completed successfully: {filename}")
        return response
        
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': f'Export failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsSurveyAdmin])
def category_analytics_pdf_export(request):
    """
    Export category analytics as PDF report with charts and statistics.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from datetime import datetime
        import io
        
        category_id = request.data.get('category_id')
        
        if not category_id:
            return Response({'error': 'category_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            category = SurveyCategory.objects.get(id=category_id)
        except SurveyCategory.DoesNotExist:
            return Response({'error': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get analytics data (reuse the analytics view logic)
        questions = SurveyQuestion.objects.filter(
            category=category,
            is_active=True
        ).order_by('order', 'question_text')
        
        responses = SurveyResponse.objects.filter(
            question__category=category
        ).select_related('user', 'question')
        
        from django.contrib.auth import get_user_model
        User = get_user_model()
        total_alumni = User.objects.filter(user_type=3, is_approved=True).count()
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        story.append(Paragraph(f"Survey Analytics Report", title_style))
        story.append(Paragraph(f"Category: {category.name}", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary section
        story.append(Paragraph("Summary", heading_style))
        summary_data = [
            ['Metric', 'Value'],
            ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Questions', str(questions.count())],
            ['Total Responses', str(responses.count())],
            ['Unique Respondents', str(responses.values('user').distinct().count())],
            ['Total Alumni', str(total_alumni)],
            ['Response Rate', f"{round((responses.values('user').distinct().count() / total_alumni * 100) if total_alumni > 0 else 0, 2)}%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Question details
        story.append(Paragraph("Question Analytics", heading_style))
        story.append(Spacer(1, 0.1*inch))
        
        for question in questions:
            question_responses = responses.filter(question=question)
            response_count = question_responses.count()
            
            # Question header
            story.append(Paragraph(f"<b>{question.question_text}</b>", styles['Normal']))
            story.append(Paragraph(f"Type: {question.get_question_type_display()} | Responses: {response_count}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
            
            # Question-specific analytics
            if question.question_type in ['radio', 'select', 'yes_no']:
                # Distribution table
                distribution = {}
                for response in question_responses:
                    value = self._extract_value_for_pdf(response.response_data)
                    if value:
                        distribution[str(value)] = distribution.get(str(value), 0) + 1
                
                if distribution:
                    dist_data = [['Option', 'Count', 'Percentage']]
                    for option, count in sorted(distribution.items(), key=lambda x: x[1], reverse=True):
                        percentage = round((count / response_count * 100) if response_count > 0 else 0, 1)
                        dist_data.append([option, str(count), f"{percentage}%"])
                    
                    dist_table = Table(dist_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
                    dist_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#60a5fa')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    story.append(dist_table)
            
            elif question.question_type == 'rating':
                # Rating statistics
                values = []
                for response in question_responses:
                    value = self._extract_value_for_pdf(response.response_data)
                    if value is not None:
                        try:
                            values.append(float(value))
                        except (ValueError, TypeError):
                            pass
                
                if values:
                    avg = round(sum(values) / len(values), 2)
                    story.append(Paragraph(f"Average Rating: <b>{avg}</b> / {question.max_value or 5}", styles['Normal']))
                    
                    # Distribution
                    distribution = {}
                    for val in values:
                        distribution[int(val)] = distribution.get(int(val), 0) + 1
                    
                    dist_data = [['Rating', 'Count']]
                    for rating in range(question.min_value or 1, (question.max_value or 5) + 1):
                        count = distribution.get(rating, 0)
                        dist_data.append([str(rating), str(count)])
                    
                    dist_table = Table(dist_data, colWidths=[2*inch, 2*inch])
                    dist_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#60a5fa')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    story.append(dist_table)
            
            elif question.question_type in ['text', 'textarea']:
                # Sample responses
                samples = []
                for response in question_responses[:5]:
                    value = self._extract_value_for_pdf(response.response_data)
                    if value:
                        samples.append(str(value)[:150])
                
                if samples:
                    story.append(Paragraph("<b>Sample Responses:</b>", styles['Normal']))
                    for i, sample in enumerate(samples, 1):
                        story.append(Paragraph(f"{i}. {sample}", styles['Normal']))
            
            story.append(Spacer(1, 0.2*inch))
        
        # Build PDF
        doc.build(story)
        
        # Return response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'survey_analytics_{category.name}_{timestamp}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError:
        return Response({
            'error': 'ReportLab library not installed. Run: pip install reportlab'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': f'PDF export failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def _extract_value_for_pdf(response_data):
    """Helper function to extract value from response_data"""
    if not response_data:
        return None
    if isinstance(response_data, dict):
        for key in ['value', 'answer', 'text', 'rating', 'selected_options']:
            if key in response_data:
                return response_data[key]
        return response_data
    return response_data


@api_view(['POST'])
@permission_classes([IsSuperAdminOnly])
def clear_survey_cache_view(request):
    """
    Clear all survey-related cache.
    Super admin only utility endpoint.
    NOTE: Registration survey no longer uses caching.
    """
    return Response({
        'message': 'No cache clearing needed - registration survey fetches live data',
        'cleared_at': timezone.now().isoformat()
    })


@method_decorator(never_cache, name='dispatch')
class RegistrationSurveyQuestionsView(APIView):
    """
    Get survey questions specifically for registration process.
    Public endpoint (no authentication required for registration).
    NO CACHING - Always fetch fresh data from database.
    """
    permission_classes = [AllowAny]
    
    def get(self, request):
        print(f"🔄 Loading fresh registration survey data from database (no cache)")
        
        # Get active categories that are marked for inclusion in registration
        # Only return categories where include_in_registration=True
        categories = SurveyCategory.objects.filter(
            is_active=True,
            include_in_registration=True
        ).prefetch_related(
            'questions'
        ).order_by('order', 'name')
        
        survey_data = []
        for category in categories:
            active_questions = category.questions.filter(is_active=True).order_by('order', 'question_text')
            
            if active_questions.exists():
                questions_data = []
                for question in active_questions:
                    question_data = {
                        'id': question.id,
                        'question_text': question.question_text,
                        'question_type': question.question_type,
                        'placeholder_text': question.placeholder_text,
                        'help_text': question.help_text,
                        'options': question.get_options_list(),
                        'is_required': question.is_required,
                        'min_value': question.min_value,
                        'max_value': question.max_value,
                        'max_length': question.max_length,
                        'order': question.order
                    }
                    
                    # Add question-level conditional logic if exists
                    if question.depends_on_question:
                        question_data['depends_on_question_id'] = question.depends_on_question.id
                        question_data['depends_on_value'] = question.depends_on_value
                    
                    questions_data.append(question_data)
                
                category_data = {
                    'category': {
                        'id': category.id,
                        'name': category.name,
                        'description': category.description,
                        'order': category.order
                    },
                    'questions': questions_data
                }
                
                # Add category-level conditional logic if exists
                if category.depends_on_category:
                    category_data['category']['depends_on_category'] = category.depends_on_category.id
                    category_data['category']['depends_on_category_name'] = category.depends_on_category.name
                    category_data['category']['depends_on_question_text'] = category.depends_on_question_text
                    category_data['category']['depends_on_value'] = category.depends_on_value
                
                survey_data.append(category_data)
        
        # Return fresh data directly (no caching)
        print(f"✅ Returning fresh survey data with {len(survey_data)} categories")
        
        response = Response(survey_data)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response


# ============================================================================
# COMPREHENSIVE ANALYTICS DASHBOARD VIEWS
# ============================================================================

class AnalyticsOverviewView(APIView):
    """
    Analytics Overview - Executive summary and KPIs
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        try:
            filters = request.data.get('filters', {})
            
            # Apply filters and get basic stats
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            users = User.objects.filter(user_type=3)  # Alumni only
            responses = SurveyResponse.objects.all()
            
            # Apply filters (implement filtering logic)
            if filters.get('programs'):
                users = users.filter(program__in=filters['programs'])
            if filters.get('graduation_years'):
                users = users.filter(year_graduated__in=filters['graduation_years'])
            
            # Calculate KPIs
            total_respondents = users.count()
            employed_count = users.filter(employment_status__in=['employed_locally', 'employed_internationally']).count()
            international_count = users.filter(employment_status='employed_internationally').count()
            
            overview_data = {
                'total_respondents': total_respondents,
                'employed_count': employed_count,
                'international_count': international_count,
                'unemployed_count': users.filter(employment_status='unemployed').count(),
                'self_employed_count': users.filter(employment_status='self_employed').count(),
                'average_income': 0,  # Calculate from survey responses
                'average_competitiveness': 0,  # Calculate from survey responses
                'studies_count': 0,  # Calculate from survey responses
                'quick_insights': [
                    f"Total of {total_respondents} alumni have responded to the survey",
                    f"Employment rate stands at {(employed_count/total_respondents*100):.1f}%" if total_respondents > 0 else "No employment data available",
                    f"International employment accounts for {(international_count/employed_count*100):.1f}% of employed alumni" if employed_count > 0 else "No international employment data"
                ]
            }
            
            return Response({'overview': overview_data})
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EmployabilityAnalyticsView(APIView):
    """
    Employment outcome analysis
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        try:
            filters = request.data.get('filters', {})
            
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            users = User.objects.filter(user_type=3)  # Alumni only
            
            # Apply filters
            if filters.get('programs'):
                users = users.filter(program__in=filters['programs'])
            
            # Calculate employment statistics
            total = users.count()
            employed = users.filter(employment_status__in=['employed_locally', 'employed_internationally']).count()
            unemployed = users.filter(employment_status='unemployed').count()
            self_employed = users.filter(employment_status='self_employed').count()
            local_employed = users.filter(employment_status='employed_locally').count()
            international_employed = users.filter(employment_status='employed_internationally').count()
            
            # Program analysis
            programs = users.values('program').distinct()
            program_analysis = []
            
            for program in programs:
                if program['program']:
                    program_users = users.filter(program=program['program'])
                    program_total = program_users.count()
                    program_employed = program_users.filter(employment_status__in=['employed_locally', 'employed_internationally']).count()
                    
                    program_analysis.append({
                        'name': program['program'],
                        'employmentRate': round((program_employed/program_total*100) if program_total > 0 else 0, 2),
                        'localEmployment': round((program_users.filter(employment_status='employed_locally').count()/program_total*100) if program_total > 0 else 0, 2),
                        'internationalEmployment': round((program_users.filter(employment_status='employed_internationally').count()/program_total*100) if program_total > 0 else 0, 2),
                        'averageIncome': 25000,  # Placeholder - calculate from actual data
                        'jobRelevancePercentage': 85,  # Placeholder - calculate from actual data
                        'jobRelevanceScore': 4.2,  # Placeholder
                        'respondentCount': program_total
                    })
            
            # Sort by employment rate
            program_analysis.sort(key=lambda x: x['employmentRate'], reverse=True)
            
            employability_data = {
                'overallStats': {
                    'employmentRate': round((employed/total*100) if total > 0 else 0, 2),
                    'unemploymentRate': round((unemployed/total*100) if total > 0 else 0, 2),
                    'selfEmployedRate': round((self_employed/total*100) if total > 0 else 0, 2),
                    'averageIncome': 25000,  # Placeholder
                    'localEmploymentRate': round((local_employed/employed*100) if employed > 0 else 0, 2),
                    'internationalRate': round((international_employed/employed*100) if employed > 0 else 0, 2)
                },
                'programAnalysis': program_analysis,
                'topPerformers': program_analysis[:5],
                'incomeAnalysis': {
                    'distribution': {'₱15,000-₱25,000': 45, '₱25,000-₱35,000': 30, '₱35,000+': 25},
                    'localAverage': 22000,
                    'internationalAverage': 45000,
                    'bySector': {'Private': 28000, 'Government': 25000, 'NGO': 22000}
                },
                'jobRelevance': {
                    'Highly Relevant': {'percentage': 65},
                    'Somewhat Relevant': {'percentage': 25},
                    'Not Relevant': {'percentage': 10}
                },
                'employmentSectors': {
                    'Private': 60,
                    'Government': 25,
                    'NGO': 10,
                    'Self-Employed': 5
                }
            }
            
            return Response({'employability': employability_data})
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SkillsAnalyticsView(APIView):
    """
    Skills relevance analysis
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        # Placeholder implementation
        skills_data = {
            'overallAverages': {
                'Communication': 4.2,
                'Critical Thinking': 4.0,
                'Leadership': 3.8,
                'Adaptability': 4.1,
                'Productivity': 3.9,
                'Entrepreneurship': 3.5
            },
            'rankedSkills': [
                {'skill': 'Communication', 'average': 4.2},
                {'skill': 'Adaptability', 'average': 4.1},
                {'skill': 'Critical Thinking', 'average': 4.0},
                {'skill': 'Productivity', 'average': 3.9},
                {'skill': 'Leadership', 'average': 3.8},
                {'skill': 'Entrepreneurship', 'average': 3.5}
            ],
            'topSkills': [
                {'skill': 'Communication', 'average': 4.2},
                {'skill': 'Adaptability', 'average': 4.1},
                {'skill': 'Critical Thinking', 'average': 4.0}
            ]
        }
        return Response({'skills': skills_data})


class CurriculumAnalyticsView(APIView):
    """
    Curriculum effectiveness analysis
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        # Placeholder implementation
        curriculum_data = {
            'componentRatings': {
                'Major Courses': 4.3,
                'OJT/Internship': 4.5,
                'General Education': 3.8,
                'Electives': 3.6,
                'Co-curricular': 3.9,
                'Extra-curricular': 3.7
            }
        }
        return Response({'curriculum': curriculum_data})


class StudiesAnalyticsView(APIView):
    """
    Further studies analysis
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        # Placeholder implementation
        studies_data = {
            'overview': {
                'participationRate': 35,
                'totalPursuing': 150
            }
        }
        return Response({'studies': studies_data})


class CompetitivenessAnalyticsView(APIView):
    """
    Competitiveness analysis
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        # Placeholder implementation
        competitiveness_data = {
            'overallScore': 4.1,
            'distribution': {'5': 30, '4': 40, '3': 20, '2': 8, '1': 2}
        }
        return Response({'competitiveness': competitiveness_data})


class ProgramComparisonAnalyticsView(APIView):
    """
    Program comparison analysis
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        # Placeholder implementation
        comparison_data = {
            'performanceMatrix': [],
            'overallRankings': []
        }
        return Response({'program_comparison': comparison_data})


class DemographicsAnalyticsView(APIView):
    """
    Demographics analysis
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        # Placeholder implementation
        demographics_data = {
            'overview': {
                'totalRespondents': 500,
                'genderDistribution': {'Male': 45, 'Female': 55},
                'civilStatusDistribution': {'Single': 60, 'Married': 35, 'Others': 5},
                'locationDistribution': {'Local': 75, 'International': 25}
            }
        }
        return Response({'demographics': demographics_data})


class AnalyticsFilterOptionsView(APIView):
    """
    Get available filter options for analytics
    """
    permission_classes = [IsSurveyAdmin]
    
    def get(self, request):
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            # Get unique programs
            programs = User.objects.filter(user_type=3, program__isnull=False).values_list('program', flat=True).distinct()
            program_options = []
            for program in programs:
                if program:
                    count = User.objects.filter(user_type=3, program=program).count()
                    program_options.append({'value': program, 'label': program, 'count': count})
            
            # Get graduation years
            years = User.objects.filter(user_type=3, year_graduated__isnull=False).values_list('year_graduated', flat=True).distinct().order_by('-year_graduated')
            year_options = []
            for year in years:
                count = User.objects.filter(user_type=3, year_graduated=year).count()
                year_options.append({'value': year, 'label': str(year), 'count': count})
            
            # Get employment statuses
            employment_options = [
                {'value': 'employed_locally', 'label': 'Employed Locally', 'count': User.objects.filter(user_type=3, employment_status='employed_locally').count()},
                {'value': 'employed_internationally', 'label': 'Employed Internationally', 'count': User.objects.filter(user_type=3, employment_status='employed_internationally').count()},
                {'value': 'self_employed', 'label': 'Self-Employed', 'count': User.objects.filter(user_type=3, employment_status='self_employed').count()},
                {'value': 'unemployed', 'label': 'Unemployed', 'count': User.objects.filter(user_type=3, employment_status='unemployed').count()},
            ]
            
            # Basic location options
            location_options = [
                {'value': 'local', 'label': 'Local (Philippines)', 'count': User.objects.filter(user_type=3, employment_status='employed_locally').count()},
                {'value': 'international', 'label': 'International', 'count': User.objects.filter(user_type=3, employment_status='employed_internationally').count()},
            ]
            
            # Gender options
            gender_options = [
                {'value': 'male', 'label': 'Male', 'count': User.objects.filter(user_type=3, sex='male').count()},
                {'value': 'female', 'label': 'Female', 'count': User.objects.filter(user_type=3, sex='female').count()},
            ]
            
            # Civil status options
            civil_status_options = [
                {'value': 'single', 'label': 'Single', 'count': User.objects.filter(user_type=3, civil_status='single').count()},
                {'value': 'married', 'label': 'Married', 'count': User.objects.filter(user_type=3, civil_status='married').count()},
                {'value': 'separated', 'label': 'Separated', 'count': User.objects.filter(user_type=3, civil_status='separated').count()},
                {'value': 'widowed', 'label': 'Widowed', 'count': User.objects.filter(user_type=3, civil_status='widowed').count()},
            ]
            
            return Response({
                'programs': program_options,
                'graduationYears': year_options,
                'employmentStatuses': employment_options,
                'locations': location_options,
                'genders': gender_options,
                'civilStatuses': civil_status_options
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AnalyticsExportView(APIView):
    """
    Export specific analytics reports
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        # Placeholder for export functionality
        return Response({'message': 'Export functionality coming soon'})


class AnalyticsFullReportView(APIView):
    """
    Export comprehensive analytics report
    """
    permission_classes = [IsSurveyAdmin]
    
    def post(self, request):
        # Placeholder for full report export
        return Response({'message': 'Full report export functionality coming soon'})
